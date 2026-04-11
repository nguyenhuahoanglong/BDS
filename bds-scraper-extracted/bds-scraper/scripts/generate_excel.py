#!/usr/bin/env python3
"""
Generate formatted Excel file from BDS listing data.

Usage:
    python generate_excel.py --data '<json_string>' --output '/path/to/output.xlsx'

Or import and call directly:
    from generate_excel import create_bds_excel
    create_bds_excel(projects, output_path, title="Chung Cu 3-5 Ty")

Data format (JSON array):
[
    {
        "name": "Topaz Elite",
        "type": "Chung cu",
        "price": "3.2-4.8 ty",
        "area": "60-92m2",
        "beds": "2-3",
        "wc": "2",
        "addr": "Cao Lo, P.4, Q.8",
        "district": "Q.8",
        "note": "Van Thai Land, 3 block",
        "link": "https://batdongsan.com.vn/...",
        "lat": 10.7392189,
        "lng": 106.6797322
    }
]
"""

import json
import sys
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# District color mapping (light backgrounds)
DISTRICT_COLORS = {
    "Q.1": "FCE4EC", "Q.3": "FFF3E0", "Q.4": "FFFDE7", "Q.5": "E8F5E9",
    "Q.6": "FFF3E0", "Q.7": "E0F2F1", "Q.8": "E8F4FD", "Q.10": "F3E5F5",
    "Q.11": "FCE4EC", "Q.12": "EFEBE9", "Go Vap": "ECEFF1",
    "Binh Thanh": "FBE9E7", "Tan Binh": "E8F5E9", "Tan Phu": "E0F7FA",
    "Phu Nhuan": "FFF3E0", "Binh Tan": "F1F8E9", "Thu Duc": "EDE7F6",
    "Cu Chi": "FAFAFA", "Hoc Mon": "F5F5F5", "Binh Chanh": "EFEBE9",
    "Nha Be": "E0F2F1", "Can Gio": "ECEFF1"
}


def get_district_color(district_str):
    """Find the matching color for a district string."""
    district_str = district_str.strip()
    for key, color in DISTRICT_COLORS.items():
        if key.lower() in district_str.lower() or district_str.lower() in key.lower():
            return color
    # Try numeric match (Q8 -> Q.8)
    import re
    m = re.search(r'(\d+)', district_str)
    if m:
        num = m.group(1)
        for key, color in DISTRICT_COLORS.items():
            if num in key:
                return color
    return None


def create_bds_excel(projects, output_path, title="BDS Listings"):
    """Create formatted Excel workbook from project data."""
    wb = Workbook()

    # ===== Sheet 1: Main listing =====
    ws1 = wb.active
    ws1.title = "Danh Sach BDS"

    # Header style
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # Headers
    headers = ["STT", "Du An", "Loai", "Khoang Gia", "Dien Tich", "Phong Ngu", "WC", "Dia Chi", "Quan", "Ghi Chu", "Link BDS"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, p in enumerate(projects, 1):
        row = i + 1
        ws1.cell(row=row, column=1, value=i).alignment = Alignment(horizontal="center")
        ws1.cell(row=row, column=2, value=p.get("name", ""))
        ws1.cell(row=row, column=3, value=p.get("type", "Chung cu"))
        ws1.cell(row=row, column=4, value=p.get("price", ""))
        ws1.cell(row=row, column=5, value=p.get("area", ""))
        ws1.cell(row=row, column=6, value=p.get("beds", "")).alignment = Alignment(horizontal="center")
        ws1.cell(row=row, column=7, value=p.get("wc", "")).alignment = Alignment(horizontal="center")
        ws1.cell(row=row, column=8, value=p.get("addr", ""))
        ws1.cell(row=row, column=9, value=p.get("district", ""))
        ws1.cell(row=row, column=10, value=p.get("note", ""))

        # Hyperlink
        link = p.get("link", "")
        if link:
            cell = ws1.cell(row=row, column=11, value="Xem BDS")
            cell.hyperlink = link
            cell.font = Font(color="0563C1", underline="single")
        else:
            ws1.cell(row=row, column=11, value="")

        # Color-code by district
        district = p.get("district", "")
        color = get_district_color(district)
        if color:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, 12):
                ws1.cell(row=row, column=col).fill = fill

        # Border all cells
        for col in range(1, 12):
            ws1.cell(row=row, column=col).border = thin_border

    # Column widths
    col_widths = [5, 22, 12, 15, 12, 10, 5, 30, 12, 25, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes and auto-filter
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:K{len(projects) + 1}"

    # ===== Sheet 2: Coordinates for Google My Maps =====
    ws2 = wb.create_sheet("Toa Do (Map)")

    map_headers = ["Du An", "Quan", "Lat", "Lng", "Gia", "Dia Chi"]
    for col, h in enumerate(map_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, p in enumerate(projects, 1):
        row = i + 1
        ws2.cell(row=row, column=1, value=p.get("name", ""))
        ws2.cell(row=row, column=2, value=p.get("district", ""))
        ws2.cell(row=row, column=3, value=p.get("lat", 0))
        ws2.cell(row=row, column=4, value=p.get("lng", 0))
        ws2.cell(row=row, column=5, value=p.get("price", ""))
        ws2.cell(row=row, column=6, value=p.get("addr", ""))
        for col in range(1, 7):
            ws2.cell(row=row, column=col).border = thin_border

    map_widths = [22, 12, 12, 12, 15, 30]
    for i, w in enumerate(map_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = "A2"

    # Save
    wb.save(output_path)
    print(f"Excel saved: {output_path} ({Path(output_path).stat().st_size:,} bytes)")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate BDS Excel file")
    parser.add_argument("--data", required=True, help="JSON string or path to JSON file with project data")
    parser.add_argument("--output", required=True, help="Output Excel file path")
    parser.add_argument("--title", default="BDS Listings", help="Sheet title")
    args = parser.parse_args()

    # Parse data
    if Path(args.data).exists():
        with open(args.data) as f:
            projects = json.load(f)
    else:
        projects = json.loads(args.data)

    create_bds_excel(projects, args.output, args.title)
